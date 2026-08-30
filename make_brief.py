#!/usr/bin/env python3
"""make_brief.py — 원천 엑셀을 읽어 제작 프롬프트(BRIEF) 초안을 채운다.

`prompts/BRIEF.md`가 형식이고, 이 스크립트는 그 다섯 칸 중 **1번(결론 후보)과
2번(원천 셀 맵)을 데이터에서 미리 채운다.** 사람은 결론 한 줄만 고르거나 고치면 된다.

왜 이게 필요한가 (2026-08-30):
  같은 데이터로 장표 다섯 장을 만들며 재작업이 어디서 나는지 셌다. 프롬프트에
  형식만 있고 결론이 없을 때 재작업이 났다. 그런데 결론은 데이터를 봐야 나온다 —
  시키는 사람이 폰에서 그걸 다 하기는 어렵다. 그래서 기계가 먼저 후보를 올린다.

**수를 지어내지 않는다.** 여기서 나오는 모든 값은 시트에서 읽은 것이고
셀 주소가 같이 붙는다. 붙지 않는 값은 내지 않는다.

  python3 make_brief.py <원천.xlsx> [-o brief.md] [--목적 "한 줄"]
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 상태 열에서 이런 말이 보이면 짚어야 할 것으로 올린다
FLAG_WORDS = ("미입금", "미수", "지연", "부족", "결품", "품절", "초과", "경고", "위험")
CONCENTRATION = 0.55   # 한 항목이 합계의 이 비율을 넘으면 "쏠림"으로 본다
SPREAD_RATIO = 2.0     # 최대/최소가 이 배수를 넘으면 "편차"로 본다


TIME_LABEL = re.compile(r"^\s*\d{4}[-/](W\d+|\d{1,2})\s*$|^\s*\d{4}[-/]\d{1,2}[-/]\d{1,2}")


def _timelike(label: str) -> bool:
    """주차·월·날짜 라벨. 이런 축의 '편차'는 대개 그냥 시간 흐름이라 소식이 아니다."""
    return bool(TIME_LABEL.match(str(label)))


def _err(v) -> bool:
    """엑셀 오류값. 라벨로도 값으로도 쓰지 않는다."""
    return isinstance(v, str) and v.startswith("#")


def data_blocks(ws) -> list[tuple[int, int, int, int]]:
    """(머리글행, 라벨열, 시작행, 끝행) 목록.

    시트 하나에 표가 여럿 있을 수 있다. 머리글을 엉뚱한 행에서 가져오면
    **이름과 수가 어긋난다** — 실제로 그 버그가 났다(2026-08-30, 월간대시보드에서
    4행 머리글에 18행 값을 붙였다). 그래서 "값 덩어리 바로 위"만 머리글로 본다.
    """
    def readable(labels: list[str]) -> float:
        """라벨이 사람이 읽을 수 있는가. 상품코드 같은 긴 숫자열이면 낮다.
        '8809729570007 하나가 합계의 60%'는 맞는 말이지만 읽는 사람에게 쓸모가 없다."""
        if not labels:
            return 0.0
        good = sum(1 for l in labels if not re.fullmatch(r"\d{6,}", l.strip()))
        return good / len(labels)

    best, best_score = [], -1.0
    for lab in range(1, min(ws.max_column, 6) + 1):
        blocks = []
        run_start = None
        for r in range(2, ws.max_row + 2):
            v = ws.cell(r, lab).value if r <= ws.max_row else None
            has_num = any(isinstance(ws.cell(r, c).value, (int, float))
                          for c in range(1, ws.max_column + 1)) if r <= ws.max_row else False
            ok = isinstance(v, str) and v.strip() and not _err(v) and has_num
            if ok and run_start is None:
                run_start = r
            elif not ok and run_start is not None:
                if r - run_start >= 3:                      # 세 줄 이상이라야 표로 본다
                    head = run_start - 1
                    texts = [ws.cell(head, c).value for c in range(1, ws.max_column + 1)]
                    if sum(1 for t in texts if isinstance(t, str) and t.strip() and not _err(t)) >= 3:
                        blocks.append((head, lab, run_start, r - 1))
                run_start = None
        if blocks:
            labels = [str(ws.cell(r, lab).value)
                      for _, _, a, b in blocks for r in range(a, b + 1)]
            score = readable(labels)
            if score > best_score:
                best, best_score = blocks, score
            if score == 1.0:                                 # 더 볼 필요 없다
                break
    return best


def findings_for(ws) -> list[tuple[float, str]]:
    """(중요도, 문장). 기계적으로 확인되는 것만 낸다. 해석은 붙이지 않는다."""
    out: list[tuple[float, str]] = []
    for head, lab, r0, r1 in data_blocks(ws):
        rows = list(range(r0, r1 + 1))
        for c in range(1, ws.max_column + 1):
            name = ws.cell(head, c).value
            if not isinstance(name, str) or not name.strip() or _err(name):
                continue
            name = name.strip()
            pairs = [(str(ws.cell(r, lab).value).strip(), ws.cell(r, c).value, ws.cell(r, c).coordinate)
                     for r in rows if isinstance(ws.cell(r, c).value, (int, float))]
            nz = [p for p in pairs if p[1] > 0]
            if len(nz) < 3:                                  # 항목이 둘 이하면 쏠림·편차가 뜻이 없다
                continue
            total = sum(v for _, v, _ in pairs)
            if total <= 0:
                continue
            top_l, top_v, top_ref = max(pairs, key=lambda p: p[1])
            share = top_v / total
            if share >= CONCENTRATION and not _timelike(top_l):
                # 쏠림이 가장 쓸모 있다. "무엇이 전체를 끌고 가나"가 바로 결론이 된다
                out.append((0.70 + share * 0.25, f"**쏠림** — `{name}`은 '{top_l}' 하나가 합계의 {share:.0%}다 "
                                   f"(`{ws.title}!{top_ref}` = {top_v:,.0f} / 합계 {total:,.0f})"))
            zeros = [(l, ref) for l, v, ref in pairs if v == 0]
            if 0 < len(zeros) <= len(pairs) / 2:
                names = ", ".join(f"'{l}'" for l, _ in zeros[:3])
                out.append((0.62, f"**0인 항목** — `{name}`이 {names}에서 0이다 "
                                 f"(`{ws.title}!{zeros[0][1]}`)"))
            lo_l, lo_v, lo_ref = min(nz, key=lambda p: p[1])
            ratio = top_v / lo_v
            # 같은 라벨끼리 비교하면 뜻이 없다. 시간 축의 편차는 대개 그냥 흐름이다
            if ratio >= SPREAD_RATIO and lo_l != top_l:
                span = 0.30 + min(ratio / 40, 0.25)
                if _timelike(lo_l) or _timelike(top_l):
                    span *= 0.5
                out.append((span,
                            f"**편차** — `{name}`이 '{lo_l}' {lo_v:,.1f}부터 '{top_l}' {top_v:,.1f}까지 "
                            f"{ratio:.1f}배 벌어진다 (`{ws.title}!{lo_ref}`~`{top_ref}`)"))

        # 상태 문자열 — 값이 0인 빈 행은 세지 않는다
        for c in range(1, ws.max_column + 1):
            hits = []
            for r in rows:
                v = ws.cell(r, c).value
                if isinstance(v, str) and any(w in v for w in FLAG_WORDS):
                    if any(isinstance(ws.cell(r, k).value, (int, float)) and ws.cell(r, k).value > 0
                           for k in range(1, ws.max_column + 1)):
                        hits.append((v.strip(), ws.cell(r, c).coordinate))
            if hits:
                head_name = ws.cell(head, c).value or "상태"
                out.append((0.95, f"**상태** — `{head_name}` 열에 '{hits[0][0]}'가 {len(hits)}건 "
                                 f"(`{ws.title}!{hits[0][1]}` 외)"))
                break
    return out


def kv_pairs(ws, limit: int = 24) -> list[str]:
    """`이름 / 바로 아래 수` 꼴의 요약 블록을 셀 주소와 함께 뽑는다."""
    out = []
    for r in range(1, min(ws.max_row, 20)):
        for c in range(1, ws.max_column + 1):
            name, below = ws.cell(r, c).value, ws.cell(r + 1, c).value
            if isinstance(name, str) and name.strip() and isinstance(below, (int, float)):
                out.append(f"  {name.strip():<22} `{ws.title}!{ws.cell(r + 1, c).coordinate}` = {below:,.0f}")
                if len(out) >= limit:
                    return out
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("source")
    ap.add_argument("-o", "--out", default="brief.md")
    ap.add_argument("--목적", dest="goal", default="")
    args = ap.parse_args()

    src = Path(args.source)
    wb = load_workbook(src, data_only=True)

    found: list[str] = []
    summary: list[str] = []
    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        found += findings_for(ws)
        kv = kv_pairs(ws)
        if kv:
            summary.append(f"**{ws.title}**\n" + "\n".join(kv))

    body = [f"# 제작 브리프 (초안) — {src.name}", "",
            "`make_brief.py`가 채운 초안이다. **1번 결론만 골라 고치면 된다.**",
            "형식 설명은 `prompts/BRIEF.md`에 있다.", "",
            "## 1. 결론 — 한 문장으로 골라 고친다", ""]
    if args.goal:
        body += [f"> {args.goal}", "", "아래는 데이터에서 기계적으로 확인된 것이다. 근거로 쓴다.", ""]
    else:
        body += ["> (여기에 한 문장. 아래 후보에서 고르거나 새로 쓴다)", ""]

    if found:
        # 폰에서 읽는다. 다 쏟아 놓으면 아무도 안 읽는다 — 중요도 순으로 여덟 개만.
        seen, ranked = set(), []
        for score, line in sorted(found, key=lambda x: -x[0]):
            if line not in seen:
                ranked.append(line)
                seen.add(line)
        body += [f"### 데이터가 스스로 말하는 것 (확인 {len(ranked)}건 중 상위)", ""]
        for line in ranked[:8]:
            body.append(f"- {line}")
        body += ["", "**해석은 붙이지 않았다.** 위는 시트에서 읽은 사실이고 셀 주소가 근거다.", ""]
    else:
        body += ["_기계적으로 잡히는 패턴이 없다. 결론을 직접 적어야 한다._", ""]

    body += ["## 2. 원천", "", f"`{src.name}`", ""]
    if summary:
        body += ["요약 수치로 보이는 것 (이름 / 셀 / 값):", "", "```"]
        body += summary
        body += ["```", ""]
    body += ["## 3. 독자", "", "> 경영진 / 실무 / 외부 — 하나 고른다", "",
             "## 4. 형식", "", "> 자유 · 표 중심 · 차트 중심 — 모르면 '자유'", "",
             "## 5. 제약 — 밝혀야 할 한계", "",
             "> 안 적으면 장표가 실제보다 좋아 보인다.",
             "> 예: 기준일이 월 중이라 마감 전 / 일부 비용 미입력 / 추정치 포함", ""]

    Path(args.out).write_text("\n".join(body), encoding="utf-8")
    print(f"초안: {args.out}  (시트 {len(wb.worksheets)}개, 확인된 것 {len(set(l for _, l in found))}건)")


if __name__ == "__main__":
    main()
